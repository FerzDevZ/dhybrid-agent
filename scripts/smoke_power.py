#!/usr/bin/env python3
"""Smoke 0.9.0 — verifikasi tool power end-to-end lewat build_tools (jalur nyata).

- 5 tool power terdaftar + allowlist default 36
- sys_info jalan sungguhan (psutil)
- data_query ke CSV sungguhan (duckdb)
- scaffold render template (jinja2)
- pdf_ops merge (pypdf)
- xlsx_edit ke salinan (openpyxl)
- read_image tolak non-gambar (magic bytes)

Keluar 0 hanya bila SEMUA lolos.
"""
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from dhybrid.config import Config
from dhybrid.tools import build_tools

POWER = {"sys_info", "scaffold", "data_query", "pdf_ops", "xlsx_edit"}


def main() -> int:
    cfg = Config.load(str(ROOT / "config" / "default.yaml"))
    reg = build_tools(cfg)

    names = {s["name"] for s in reg.specs()}
    missing = POWER - names
    assert not missing, f"tool power tidak terdaftar: {missing}"
    for n in POWER:
        assert n in reg.allowlist, f"{n} tidak di allowlist"
    assert len(reg.allowlist) == 36, f"allowlist = {len(reg.allowlist)}, harus 36"
    print(f"[OK] 5 tool power terdaftar + allowlist {len(reg.allowlist)}")

    out = reg.execute("sys_info", {})
    assert "CPU" in out and "RAM" in out, out
    print(f"[OK] sys_info: {out.splitlines()[0]}")

    with tempfile.TemporaryDirectory() as td:
        d = Path(td)
        csv = d / "data.csv"
        csv.write_text("nama,umur\nAndi,20\nBudi,30\n")
        out = reg.execute("data_query", {"sql": f"SELECT * FROM read_csv_auto('{csv}') WHERE umur > 25"})
        assert "Budi" in out and "Andi" not in out, out
        print(f"[OK] data_query: {out.splitlines()[1]}")

        tmpl = d / "tmpl"
        tmpl.mkdir()
        (tmpl / "halo.txt.j2").write_text("Halo {{ nama }}!")
        out = reg.execute("scaffold", {"template_dir": str(tmpl), "target_dir": str(d / "out"), "variables": {"nama": "Dunia"}})
        assert "OK" in out and (d / "out" / "halo.txt").read_text() == "Halo Dunia!", out
        print(f"[OK] scaffold: {out}")

        from pypdf import PdfWriter

        pdfs = []
        for i in (1, 2):
            w = PdfWriter()
            w.add_blank_page(200, 200)
            p = d / f"p{i}.pdf"
            with open(p, "wb") as fh:
                w.write(fh)
            pdfs.append(str(p))
        out = reg.execute("pdf_ops", {"sources": pdfs, "target": str(d / "merged.pdf")})
        assert "OK" in out and (d / "merged.pdf").exists(), out
        print(f"[OK] pdf_ops: {out}")

        from openpyxl import Workbook, load_workbook

        wb = Workbook()
        wb.active["A1"] = "lama"
        wb.save(d / "in.xlsx")
        out = reg.execute("xlsx_edit", {"source": str(d / "in.xlsx"), "target": str(d / "out.xlsx"), "edits": [{"cell": "A1", "value": "baru"}]})
        assert "OK" in out and load_workbook(d / "out.xlsx").active["A1"].value == "baru", out
        assert load_workbook(d / "in.xlsx").active["A1"].value == "lama"
        print(f"[OK] xlsx_edit: {out}")

    with tempfile.TemporaryDirectory() as td2:
        fake_png = Path(td2) / "fake.png"
        fake_png.write_text("ini teks, bukan gambar")
        out = reg.execute("read_image", {"path": str(fake_png)})
        assert "bukan gambar" in out, out
    print(f"[OK] read_image tolak non-gambar: {out[:60]}...")

    print("\n[SMOKE OK] power tools 0.9.0 berfungsi end-to-end")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except AssertionError as e:
        print(f"\n[SMOKE GAGAL] {e}", file=sys.stderr)
        sys.exit(1)
