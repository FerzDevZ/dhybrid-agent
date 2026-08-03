"""Tool power: xlsx_edit — edit Excel via openpyxl (tulis ke file SALINAN).

File asli tidak pernah diubah; hasil edit ditulis ke `target` baru.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _xlsx_edit(source: str, target: str, edits: list[dict]) -> str:
    if not Path(source).exists():
        return f"ERROR: file tidak ada: {source}"
    wb = load_workbook(source)
    ws = wb.active
    n = 0
    for e in edits or []:
        if "cell" in e:
            ws[e["cell"]] = e.get("value")
            n += 1
        elif "append" in e:
            ws.append(e["append"])
            n += 1
    Path(target).parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return f"OK: {n} edit diterapkan → {target} (file asli tidak diubah)"


def _default_need(reg, name, mods, description, parameters, fn) -> None:
    reg.register(name, description, parameters, fn)


def register(reg, _need=None, **kw) -> None:
    """Daftarkan xlsx_edit; _need dipakai soft.py untuk soft-register."""
    (_need or _default_need)(
        reg,
        "xlsx_edit",
        ["openpyxl"],
        "Edit Excel: set cell atau append baris; hasil ditulis ke file SALINAN (asli aman)",
        {
            "source": {"type": "string"},
            "target": {"type": "string"},
            "edits": {"type": "array"},
        },
        _xlsx_edit,
    )
