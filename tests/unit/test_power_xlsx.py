"""Task 7: tool xlsx_edit — edit Excel via openpyxl (salinan, asli aman)."""
from dhybrid.tools import power_xlsx


def _mk_xlsx(tmp_path, name):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "lama"
    wb.save(tmp_path / name)
    return str(tmp_path / name)


def test_xlsx_set_cell(tmp_path):
    from openpyxl import load_workbook

    src = _mk_xlsx(tmp_path, "in.xlsx")
    out = power_xlsx._xlsx_edit(
        src, str(tmp_path / "out.xlsx"), [{"cell": "A1", "value": "baru"}]
    )
    assert "OK" in out
    assert load_workbook(tmp_path / "out.xlsx").active["A1"].value == "baru"
    # file asli tidak berubah
    assert load_workbook(src).active["A1"].value == "lama"


def test_xlsx_append_row(tmp_path):
    from openpyxl import load_workbook

    src = _mk_xlsx(tmp_path, "in2.xlsx")
    power_xlsx._xlsx_edit(src, str(tmp_path / "out2.xlsx"), [{"append": ["Andi", 20]}])
    ws = load_workbook(tmp_path / "out2.xlsx").active
    assert ws["A2"].value == "Andi" and ws["B2"].value == 20


def test_xlsx_blocks_missing_file(tmp_path):
    out = power_xlsx._xlsx_edit(
        str(tmp_path / "no.xlsx"), str(tmp_path / "o.xlsx"), [{"cell": "A1", "value": "x"}]
    )
    assert "ERROR" in out
